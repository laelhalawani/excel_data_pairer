from typing import List, Optional, Any, Dict
from pydantic import BaseModel, Field, ValidationError
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import json
import os

# -------------------------------
# Pydantic Models
# -------------------------------

class CellRange(BaseModel):
    """
    Represents a range of cells within a column and specified rows in an Excel sheet.
    """
    column: Optional[str] = Field(
        None,
        description="The column identifier (e.g., 'A', 'B', etc.)."
    )
    rows_range: Optional[str] = Field(
        None,
        description="The range of rows (e.g., '1-10')."
    )

class DataPair(BaseModel):
    """
    Represents a pair of source and target (MT) cell ranges within a sheet.
    """
    src: CellRange = Field(
        default_factory=CellRange,
        description="Source cell range."
    )
    mt: CellRange = Field(
        default_factory=CellRange,
        description="MT cell range."
    )

class SheetSchema(BaseModel):
    """
    Represents a single sheet within the Excel file, identified by sheet_id and containing data pairs.
    """
    sheet_id: Optional[str] = Field(
        None,
        description="Unique identifier for the sheet."
    )
    sheet_data: List[DataPair] = Field(
        default_factory=list,
        description="List of data pairs within the sheet."
    )

class FileSchema(BaseModel):
    """
    Represents the overall schema of the Excel file, including its path and contained sheets.
    """
    file_path: str = Field(
        ...,
        description="Path to the Excel file."
    )
    file_data: List[SheetSchema] = Field(
        default_factory=list,
        description="List of sheets in the Excel file."
    )

# -------------------------------
# ExcelNavigator Class
# -------------------------------

class ExcelNavigator:
    """
    A class to navigate and manipulate Excel files based on a defined schema using Pydantic and OpenPyXL.
    
    Features:
        - Add and manage sheets within the Excel schema.
        - Add and manage data pairs within sheets.
        - Retrieve data from specified ranges using OpenPyXL.
        - Serialize and deserialize the schema to/from JSON.
        - List available sheets in the Excel file.
        - Preview data from a specified range in a sheet.
        - Autosave and autoload configuration.
        - List and select Excel files from a directory.
    """

    def __init__(self, file_path: Optional[str] = None, autoload: bool = False):
        """
        Initialize the ExcelNavigator. If file_path is provided, loads the Excel file.
        
        Args:
            file_path (Optional[str]): Path to the Excel file.
            autoload (bool): If True, attempts to load configuration from an autosave file.
        """
        self.file_schema: Optional[FileSchema] = None
        self.autosave = False  # Autosave is disabled until a file is selected
        self.waiting_for_autosave = autoload
        self.autosave_dir = './autosaves'
        os.makedirs(self.autosave_dir, exist_ok=True)
        self.autosave_path: Optional[str] = None
        self.workbook = None
        if file_path:
            self.select_excel_file(file_path=file_path, autoload=autoload)

    # ---------------------------
    # File Selection Methods
    # ---------------------------

    @staticmethod
    def list_excel_files(dir_path: str) -> List[str]:
        """
        List all Excel files in the specified directory.
        
        Args:
            dir_path (str): Path to the directory to search for Excel files.
        
        Returns:
            List[str]: List of Excel file names in the directory.
        
        Raises:
            FileNotFoundError: If the directory does not exist.
        """
        excel_extensions = ('.xlsx', '.xlsm', '.xltx', '.xltm')
        try:
            files = os.listdir(dir_path)
            excel_files = [f for f in files if f.endswith(excel_extensions) and os.path.isfile(os.path.join(dir_path, f))]
            return excel_files
        except FileNotFoundError:
            raise FileNotFoundError(f"The directory '{dir_path}' does not exist.")

    def select_excel_file(
        self, 
        file_path: Optional[str] = None, 
        dir_path: Optional[str] = None,
        file_id: Optional[int] = None, 
        file_name: Optional[str] = None, 
        autoload: Optional[bool] = None,
    ) -> None:
        """
        Select an Excel file to work with.
        
        Args:
            file_path (Optional[str]): Full path to the Excel file.
            dir_path (Optional[str]): Directory path to search for files.
            file_id (Optional[int]): Index of the file in the directory listing.
            file_name (Optional[str]): Name of the Excel file in the directory.
            autoload (bool): If True, attempts to load configuration from an autosave file.
        
        Raises:
            ValueError: If file cannot be found or parameters are invalid.
        """
        # Determine whether to enable autosave based on parameters and existing state
        if autoload is True or (autoload is None and self.waiting_for_autosave):
            self.waiting_for_autosave = False
            autoload = True
            self.autosave = True
        elif autoload is False:
            self.waiting_for_autosave = False
            self.autosave = False

        # Determine file_path based on provided parameters
        if file_path:
            if not os.path.isfile(file_path):
                raise FileNotFoundError(f"The file '{file_path}' does not exist.")
        elif dir_path:
            excel_files = self.list_excel_files(dir_path)
            if not excel_files:
                raise FileNotFoundError(f"No Excel files found in directory '{dir_path}'.")
            if file_id is not None:
                try:
                    file_name = excel_files[file_id]
                except IndexError:
                    raise ValueError(f"No file found with file_id {file_id} in directory '{dir_path}'.")
            elif file_name:
                if file_name not in excel_files:
                    raise FileNotFoundError(f"The file '{file_name}' does not exist in directory '{dir_path}'.")
            else:
                raise ValueError("Either 'file_id' or 'file_name' must be provided when 'dir_path' is specified.")
            file_path = os.path.join(dir_path, file_name)
        else:
            raise ValueError("Must provide 'file_path' or 'dir_path' with 'file_id' or 'file_name'.")

        # Now, initialize the file schema, workbook, autosave path, etc.
        self.file_schema = FileSchema(file_path=file_path)
        self.autosave_path = os.path.join(
            self.autosave_dir, 
            os.path.splitext(os.path.basename(file_path))[0] + "_autosave.json"
        )
        try:
            self.workbook = load_workbook(filename=file_path, data_only=True)
            print(f"Excel file '{file_path}' loaded successfully.")
        except Exception as e:
            raise ValueError(f"Failed to load Excel file '{file_path}': {e}")
        
        if autoload:
            self._autoload_config()

    # ---------------------------
    # Autosave and Autoload Methods
    # ---------------------------

    def enable_autosave(self):
        """
        Enable automatic saving of the configuration to a JSON file.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file before enabling autosave.")
        self.autosave = True
        print("Autosave enabled.")

    def disable_autosave(self):
        """
        Disable automatic saving of the configuration.
        """
        self.autosave = False
        print("Autosave disabled.")

    def _autosave_config(self):
        """
        Automatically save the current schema to an autosave file if autosave is enabled.
        """
        if self.autosave and self.file_schema:
            try:
                with open(self.autosave_path, 'w', encoding='utf-8') as f:
                    f.write(self.to_json())
                print(f"Autosaved schema to '{self.autosave_path}'.")
            except Exception as e:
                print(f"Failed to autosave schema: {e}")

    def _autoload_config(self):
        """
        Automatically load the schema from an autosave file if it exists.
        """
        if os.path.exists(self.autosave_path):
            try:
                self.load_from_file(self.autosave_path)
                print(f"Configuration autoloaded from '{self.autosave_path}'.")
            except Exception as e:
                print(f"Failed to autoload configuration: {e}")
        else:
            print("No autosave file found to autoload.")

    # ---------------------------
    # Sheet Management Methods
    # ---------------------------

    def add_sheet(self, sheet_id: str, present_ok:bool=True) -> None:
        """
        Add a new sheet to the Excel schema.
        
        Args:
            sheet_id (str): Unique identifier for the sheet.
        
        Raises:
            ValueError: If a sheet with the given sheet_id already exists or no file is selected.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file before adding a sheet.")
        if self._find_sheet(sheet_id):
            if not present_ok:
                raise ValueError(f"Sheet with id '{sheet_id}' already exists in the schema.")
            else:
                print(f"Sheet with id '{sheet_id}' already exists in the schema.")
                return
        if sheet_id not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_id}' does not exist in the Excel file.")
        new_sheet = SheetSchema(sheet_id=sheet_id)
        self.file_schema.file_data.append(new_sheet)
        print(f"Sheet '{sheet_id}' added to schema successfully.")
        self._autosave_config()

    def remove_sheet(self, sheet_id: str) -> None:
        """
        Remove a sheet from the schema and the workbook.
        
        Args:
            sheet_id (str): Identifier of the sheet to remove.
        
        Raises:
            ValueError: If the sheet does not exist in the schema or workbook.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file before removing a sheet.")
        sheet_schema = self._find_sheet(sheet_id)
        if not sheet_schema:
            raise ValueError(f"Sheet with id '{sheet_id}' does not exist in the schema.")
        if sheet_id not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_id}' does not exist in the Excel file.")

        # Remove from schema
        self.file_schema.file_data.remove(sheet_schema)

        # Remove from workbook
        del self.workbook[sheet_id]
        try:
            self.workbook.save(self.file_schema.file_path)
            print(f"Sheet '{sheet_id}' removed successfully from schema and workbook.")
            self._autosave_config()
        except Exception as e:
            print(f"Failed to save workbook after removing sheet '{sheet_id}': {e}")

    def list_sheets(self) -> List[str]:
        """
        List all sheet identifiers in the schema.
        
        Returns:
            List[str]: List of sheet IDs.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file to list sheets.")
        return [sheet.sheet_id for sheet in self.file_schema.file_data if sheet.sheet_id]

    # ---------------------------
    # Data Pair Management Methods
    # ---------------------------

    def add_data_pair(
        self, 
        sheet_id: str, 
        src_column: str, 
        src_rows: str, 
        mt_column: str, 
        mt_rows: str,
        present_ok:bool=True
    ) -> None:
        """
        Add a data pair to a specified sheet in the Excel schema.
        
        Args:
            sheet_id (str): Identifier of the sheet to add the data pair to.
            src_column (str): Source column identifier (e.g., 'A').
            src_rows (str): Source rows range (e.g., '1-10').
            mt_column (str): MT column identifier (e.g., 'B').
            mt_rows (str): MT rows range (e.g., '1-10').
        
        Raises:
            ValueError: If the specified sheet does not exist or a duplicate data pair is detected.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file before adding a data pair.")
        sheet = self._find_sheet(sheet_id)
        if not sheet:
            raise ValueError(f"Sheet with id '{sheet_id}' does not exist in the schema.")

        # Create the new DataPair
        new_data_pair = DataPair(
            src=CellRange(column=src_column, rows_range=src_rows),
            mt=CellRange(column=mt_column, rows_range=mt_rows)
        )

        # Check for duplicates
        for existing_pair in sheet.sheet_data:
            if existing_pair == new_data_pair:
                if not present_ok:
                    raise ValueError(f"Duplicate DataPair detected in sheet '{sheet_id}'. The DataPair with src={new_data_pair.src} and mt={new_data_pair.mt} already exists.")
                else:
                    print(f"The Data Pair with src={new_data_pair.src} and mt={new_data_pair.mt} already exists. Skipping.")
        # If no duplicate, add the new DataPair
        sheet.sheet_data.append(new_data_pair)
        print(f"Data pair added to sheet '{sheet_id}' successfully.")
        self._autosave_config()


    def remove_data_pair(self, sheet_id: str, index: int) -> None:
        """
        Remove a data pair from a sheet by its index in the schema.
        
        Args:
            sheet_id (str): Identifier of the sheet.
            index (int): Index of the data pair to remove.
        
        Raises:
            ValueError: If the sheet or data pair does not exist.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file before removing a data pair.")
        sheet_schema = self._find_sheet(sheet_id)
        if not sheet_schema:
            raise ValueError(f"Sheet with id '{sheet_id}' does not exist in the schema.")
        try:
            sheet_schema.sheet_data.pop(index)
            print(f"Removed data pair at index {index} from sheet '{sheet_id}'.")
            self._autosave_config()
        except IndexError:
            raise ValueError(f"Data pair index '{index}' is out of range for sheet '{sheet_id}'.")

    def list_data_pairs(self, sheet_id: str) -> List[DataPair]:
        """
        List all data pairs in a specified sheet.
        
        Args:
            sheet_id (str): Identifier of the sheet.
        
        Returns:
            List[DataPair]: List of data pairs.
        
        Raises:
            ValueError: If the sheet does not exist.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file to list data pairs.")
        sheet = self._find_sheet(sheet_id)
        if not sheet:
            raise ValueError(f"Sheet with id '{sheet_id}' does not exist in the schema.")
        return sheet.sheet_data

    # ---------------------------
    # Data Retrieval Methods
    # ---------------------------

    def get_data(self, sheet_id: str, src: CellRange, mt: CellRange) -> Dict[str, List[Any]]:
        """
        Retrieve data from specified source and MT ranges within a sheet.
        
        Args:
            sheet_id (str): Identifier of the sheet.
            src (CellRange): Source cell range.
            mt (CellRange): MT cell range.
        
        Returns:
            Dict[str, List[Any]]: A dictionary containing data from src and mt ranges.
        
        Raises:
            ValueError: If the sheet does not exist or ranges are invalid.
        """
        if not self.file_schema or not self.workbook:
            raise ValueError("No Excel file selected. Please select an Excel file before retrieving data.")
        sheet_schema = self._find_sheet(sheet_id)
        if not sheet_schema:
            raise ValueError(f"Sheet with id '{sheet_id}' does not exist in the schema.")
        
        if sheet_id not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_id}' does not exist in the Excel file.")
        
        sheet = self.workbook[sheet_id]
        
        src_data = self._read_range(sheet, src)
        mt_data = self._read_range(sheet, mt)
        
        return {
            "src": src_data,
            "mt": mt_data
        }

    def get_all_data(self) -> Dict[str, List[Dict[str, List[Any]]]]:
        """
        Retrieve all data pairs across all sheets in the schema.
        
        Returns:
            Dict[str, List[Dict[str, List[Any]]]]: A dictionary with sheet IDs as keys and lists of data pairs as values.
        """
        if not self.file_schema or not self.workbook:
            raise ValueError("No Excel file selected. Please select an Excel file before retrieving data.")
        all_data = {}
        for sheet_schema in self.file_schema.file_data:
            sheet_id = sheet_schema.sheet_id
            if not sheet_id:
                continue  # Skip sheets without an ID
            if sheet_id not in self.workbook.sheetnames:
                print(f"Warning: Sheet '{sheet_id}' is defined in schema but does not exist in the Excel file.")
                continue
            sheet = self.workbook[sheet_id]
            all_data[sheet_id] = []
            for data_pair in sheet_schema.sheet_data:
                try:
                    data = self.get_data(sheet_id, data_pair.src, data_pair.mt)
                    all_data[sheet_id].append(data)
                except ValueError as ve:
                    print(f"Error retrieving data from sheet '{sheet_id}': {ve}")
        return all_data

    def preview_range(self, sheet_id: str, column: str, rows_range: str) -> List[Any]:
        """
        Output a list of values from a specified range in a given sheet.
        
        Args:
            sheet_id (str): Identifier of the sheet.
            column (str): Column identifier (e.g., 'A').
            rows_range (str): Rows range (e.g., '1-10').
        
        Returns:
            List[Any]: A list of cell values from the specified range.
        
        Raises:
            ValueError: If the sheet does not exist or the range is invalid.
        """
        if not self.workbook:
            raise ValueError("No Excel file selected. Please select an Excel file before previewing data.")
        if sheet_id not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_id}' does not exist in the Excel file.")

        sheet = self.workbook[sheet_id]
        cell_range = CellRange(column=column, rows_range=rows_range)
        data = self._read_range(sheet, cell_range)
        return data

    def list_file_sheets(self) -> List[str]:
        """
        List all available sheet names in the Excel file.
        
        Returns:
            List[str]: List of sheet names in the Excel workbook.
        
        Raises:
            ValueError: If no file is selected.
        """
        if not self.workbook:
            raise ValueError("No Excel file selected. Please select an Excel file to list its sheets.")
        return self.workbook.sheetnames

    # ---------------------------
    # Helper Methods
    # ---------------------------

    def _find_sheet(self, sheet_id: str) -> Optional[SheetSchema]:
        """
        Find a sheet in the schema by its identifier.
        
        Args:
            sheet_id (str): Identifier of the sheet.
        
        Returns:
            Optional[SheetSchema]: The sheet schema if found, else None.
        """
        if not self.file_schema:
            return None
        for sheet in self.file_schema.file_data:
            if sheet.sheet_id == sheet_id:
                return sheet
        return None

    def _read_range(self, sheet, cell_range: CellRange) -> List[Any]:
        """
        Read data from a specified cell range within a sheet.
        
        Args:
            sheet: OpenPyXL worksheet object.
            cell_range (CellRange): The cell range to read.
        
        Returns:
            List[Any]: A list of cell values.
        
        Raises:
            ValueError: If column or row range is invalid.
        """
        if not cell_range.column or not cell_range.rows_range:
            raise ValueError("Both 'column' and 'rows_range' must be specified in CellRange.")
        
        # Parse the column
        try:
            start_col = column_index_from_string(cell_range.column.upper())
            end_col = start_col  # Single column as per schema
        except ValueError:
            raise ValueError(f"Invalid column identifier: '{cell_range.column}'.")
        
        if '-' not in cell_range.rows_range:
            raise ValueError(f"Invalid rows_range `{cell_range.rows_range}` format: missing '-', for a single row use '1-1'.")
        # Parse the row range
        try:
            row_start, row_end = map(int, cell_range.rows_range.split('-'))
            if row_start > row_end:
                raise ValueError(f"Invalid rows_range: start row {row_start} is greater than end row {row_end}.")
        except Exception:
            raise ValueError(f"Invalid rows_range format: '{cell_range.rows_range}'. Expected format 'start-end'.")
        
        # Collect data
        data = []
        for row in sheet.iter_rows(min_row=row_start, max_row=row_end, min_col=start_col, max_col=end_col):
            for cell in row:
                data.append(cell.value)
        
        return data

    # ---------------------------
    # Schema Serialization Methods
    # ---------------------------

    def get_schema(self) -> FileSchema:
        """
        Retrieve the current file schema.
        
        Returns:
            FileSchema: The current file schema.
        
        Raises:
            ValueError: If no file is selected.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file to get the schema.")
        return self.file_schema

    def to_json(self, indent: int = 4) -> str:
        """
        Serialize the file schema to a JSON-formatted string.
        
        Args:
            indent (int, optional): Indentation level for JSON formatting. Defaults to 4.
        
        Returns:
            str: JSON representation of the file schema.
        
        Raises:
            ValueError: If no file is selected.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file to serialize the schema.")
        # Use Pydantic v2's model_dump_json
        return self.file_schema.model_dump_json(indent=indent)

    def save_to_file(self, output_path: Optional[str] = None) -> None:
        """
        Save the current schema to a JSON file.
        
        Args:
            output_path (Optional[str], optional): Path to save the JSON file. 
                                                  Defaults to the file_path with a .json extension.
        
        Raises:
            ValueError: If no file is selected.
        """
        if not self.file_schema:
            raise ValueError("No Excel file selected. Please select an Excel file before saving the schema.")
        if output_path is None:
            output_path = os.path.splitext(self.file_schema.file_path)[0] + ".json"
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(self.to_json())
            print(f"Schema saved to '{output_path}' successfully.")
            self._autosave_config()
        except Exception as e:
            print(f"Failed to save schema to '{output_path}': {e}")

    def load_from_file(self, json_path: str) -> None:
        """
        Load the schema from a JSON file.
        
        Args:
            json_path (str): Path to the JSON file.
        
        Raises:
            FileNotFoundError: If the JSON file does not exist.
            ValidationError: If the JSON does not conform to the schema.
            ValueError: If loading the Excel file fails.
        """
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"The JSON file '{json_path}' does not exist.")
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                json_data = f.read()
            # Use Pydantic v2's model_validate_json
            self.file_schema = FileSchema.model_validate_json(json_data)
            # Reload the workbook to match the updated schema
            self.workbook = load_workbook(filename=self.file_schema.file_path, data_only=True)
            print(f"Schema loaded from '{json_path}' successfully.")
            self._autosave_config()
        except ValidationError as e:
            raise ValueError(f"Schema validation failed: {e}")
        except Exception as e:
            raise ValueError(f"Failed to load schema from '{json_path}': {e}")

    # ---------------------------
    # Workbook Manipulation Methods
    # ---------------------------

    def update_cell(self, sheet_id: str, cell: str, value: Any) -> None:
        """
        Update the value of a specific cell in a sheet and save the workbook.
        
        Args:
            sheet_id (str): Identifier of the sheet.
            cell (str): Cell identifier (e.g., 'A1').
            value (Any): New value for the cell.
        
        Raises:
            ValueError: If the sheet does not exist or no file is selected.
        """
        if not self.workbook:
            raise ValueError("No Excel file selected. Please select an Excel file before updating cells.")
        if sheet_id not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_id}' does not exist in the Excel file.")
        
        sheet = self.workbook[sheet_id]
        sheet[cell] = value
        try:
            self.workbook.save(self.file_schema.file_path)
            print(f"Cell '{cell}' in sheet '{sheet_id}' updated to '{value}'.")
            self._autosave_config()
        except Exception as e:
            print(f"Failed to save workbook after updating cell '{cell}': {e}")

    # -------------------------------
    # End of ExcelNavigator Class
    # -------------------------------
